from openpyxl import load_workbook

from openpyxl.utils import get_column_letter

from collections import defaultdict

import cx_Oracle

import sys

import pprint

from pathlib import Path


def def_value():
    return None


#old_site = 'BEDFORD'
#old_org = 'EAGLENA'
new_site = 'NJ'
new_org = 'JOHNSON'
sql_folder_name = 'SQL_STG2MX'

sequence_increment = 1000
max_schema = 'maximo'
legacy_attr_prefix = 'LGCY_'

#DB Connection Details
db_user = 'stage'
db_pw = 'stage'
db_connect = 'localhost:1521/max1'


#obj_level = 'ORG'

obj_level_list = ['SITE', 'ORG', 'ORGSITE', 'SYSTEMORG', 'SYSTEMORGSITE']

for obj_level in obj_level_list:
    folder_loc = f"./{sql_folder_name}/{obj_level}"
    object_per_file = 10
    Path(folder_loc).mkdir(parents=True, exist_ok=True)
    def_filter = '1=1'

    if obj_level == 'SITE':
        def_filter += " and siteid='" + new_site + "'"

    if obj_level == 'ORG' or obj_level == 'SYSTEMORG':
        def_filter += " and orgid='" + new_org + "'"

    if obj_level == 'ORGSITE' or obj_level == 'SYSTEMORGSITE':
        def_filter += " and (siteid='" + new_site + \
            "' or orgid='" + new_org + "')"

    ATTR_TYPE = 'type'

    ATTR_SEQ = 'sequence'

    SAME_AS_ATTR = 'same_as_attribute'

    mx_Obj_attr = defaultdict(def_value)

    type_map = {}
    exclude_map = {}

    wb = load_workbook('./MX_META.xlsx')

    # print(wb.sheetnames)

    ws_mxmeta = wb[f'meta-{obj_level.lower()}']

    ws_map = wb['map']

    ws_excluded = wb['excluded']

    # print(ws_mxmeta['A1'].value,ws_mxmeta['B1'].value)

    # Load Map

    # first row is for header

    current_row = 1

    while True:

        current_row += 1

        mx_type = ws_map[get_column_letter(1) + str(current_row)].value

        sql_type = ws_map[get_column_letter(2) + str(current_row)].value

        if not mx_type:

            break

        type_map[mx_type] = sql_type

    # print(type_map)

    # first row is for header

    current_row = 1

    while True:

        current_row += 1

        mx_obj = ws_mxmeta[get_column_letter(1) + str(current_row)].value

        mx_attr = ws_mxmeta[get_column_letter(2) + str(current_row)].value

        mx_type = ws_mxmeta[get_column_letter(3) + str(current_row)].value

        mx_seq = ws_mxmeta[get_column_letter(4) + str(current_row)].value

        mx_same_as_attr = ws_mxmeta[get_column_letter(
            5) + str(current_row)].value

        if not mx_obj:

            break

        obj_dict = mx_Obj_attr.get(mx_obj)

        if obj_dict:

            obj_dict[mx_attr] = {
                ATTR_TYPE: mx_type,
                ATTR_SEQ: mx_seq,
                SAME_AS_ATTR: mx_same_as_attr}

        else:

            mx_Obj_attr[mx_obj] = {
                mx_attr: {
                    ATTR_TYPE: mx_type,
                    ATTR_SEQ: mx_seq,
                    SAME_AS_ATTR: mx_same_as_attr}}

        # print(ws_mxmeta[get_column_letter(1)+str(current_row)].value,

        #      ws_mxmeta[get_column_letter(2)+str(current_row)].value,

        #      ws_mxmeta[get_column_letter(3)+str(current_row)].value,

        #      ws_mxmeta[get_column_letter(4)+str(current_row)].value)

    current_row = 1

    while True:

        current_row += 1

        mx_obj = ws_excluded[get_column_letter(1) + str(current_row)].value

        mx_attr = ws_excluded[get_column_letter(2) + str(current_row)].value

        if not mx_obj:

            break

        if mx_attr and mx_Obj_attr.get(
                mx_obj) and mx_Obj_attr[mx_obj].get(mx_attr):
            del mx_Obj_attr[mx_obj][mx_attr]
        elif mx_Obj_attr.get(mx_obj):
            del mx_Obj_attr[mx_obj]

    # print(mx_Obj_attr)

    mx_obj_list = mx_Obj_attr.keys()
    sql_file = None
    toc_file = None
    toc_content = {}
    con = None
    object_count = 0
    try:

        #sql_file = open(folder_loc,"w")
        con = cx_Oracle.connect(
            db_user, db_pw, db_connect, encoding="UTF-8")
        cur = con.cursor()

        for dict_object in mx_obj_list:

            file_counter = int(object_count / object_per_file) + 1
            object_count += 1
            if (object_count % object_per_file == 1):
                current_file_name = f'{obj_level}_{file_counter}'
                if sql_file:
                    sql_file.close()
                sql_file = open(f'{folder_loc}/{current_file_name}.sql', "w")
                sql_file.write('Set define off;\n\n')
                sql_file.write('-' * 50)
                toc_content[file_counter] = []

            toc_content[file_counter].append(dict_object)
            # print(dict_object)

            #current_file_name = obj_level

            dict_atrr = mx_Obj_attr[dict_object]

            list_attr = list(dict_atrr.keys())

            attr_type_dict = {}

            attr_seq_dict = {}

            attr_siteorg = {}

            for attr in list_attr:

                type = dict_atrr[attr][ATTR_TYPE]

                sequence = dict_atrr[attr][ATTR_SEQ]

                same_as_attr = dict_atrr[attr][SAME_AS_ATTR]

                attr_type_dict[attr] = type

                if sequence:

                    attr_seq_dict[attr] = sequence

                if same_as_attr and (
                        same_as_attr == 'ORGID' or same_as_attr == 'SITEID'):

                    attr_siteorg[attr] = same_as_attr

            # print(attr_type_dict)

            # print(attr_seq_dict)

            select_statement = f"SELECT {','.join(list_attr)} FROM {dict_object} where " + \
                def_filter

            master_insert = f"INSERT INTO {dict_object} ({','.join(list_attr)}) values()"

            # print(select_statement)

            sql_file.write(dict_object)
            sql_file.write('-' * 50)
            sql_file.write("\n\n")
            
            
            for result in cur.execute(select_statement):
                
                #legacy_values = {}
                val_list = []

                index = 0

                for item in result:

                    mapped_value = None

                    if item is None:

                        mapped_value = 'NULL'
                    # print('processing......')

                    select_attr = list_attr[index]

                    #print("selected attribute is",select_attr)
                    
                        


                    select_attr_type = attr_type_dict.get(select_attr)

                    #print("selected attribute type is",select_attr_type)

                    if not mapped_value and select_attr_type:

                        mapped_attr_type = type_map.get(select_attr_type)
                        #print("Mapped attribute type",mapped_attr_type)

                        if mapped_attr_type == 'NUM' or mapped_attr_type == 'FLOAT':

                            mapped_value = str(item)

                        elif mapped_attr_type == 'ALN':

                            mapped_value = repr(item)

                        elif mapped_attr_type == 'DATETIME':

                            mapped_value = "TO_DATE(" + repr(str(item)) + \
                                ",'YYYY-MM-DD HH24:MI:SS')"

                    val_list.append(mapped_value)

                    index += 1

                # print(val_list)
                

                master_insert = f"INSERT INTO {dict_object} ({','.join(list_attr)}) values({','.join(val_list)});\n"

                # print(master_insert)
                sql_file.write(master_insert)

            sql_file.write("\n\n")
            sql_file.write("commit;")
            sql_file.write("\n\n")
            sql_file.write('-' * 50)

        toc_file = sql_file = open(folder_loc + "/Index.txt", "w")
        toc_file.write(pprint.pformat(toc_content))
        toc_file.close()

    except Exception as e:

        _, __, exc_tb = sys.exc_info()

        print("Exception raised in DB Block", e,
              "Line Number is", exc_tb.tb_lineno)

    finally:

        if con:
            con.close()
            print("Finally closing connection")
        if sql_file:
            sql_file.close()
        if toc_file:
            toc_file.close()

        #master_insert = f"INSERT INTO {dict_object} ({','.join(list_attr)}) values()"

        # print(master_insert)

    wb.close()
    # Done 1.1
